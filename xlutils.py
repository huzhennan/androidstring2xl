
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook, Worksheet
import os
from re import match
from android2po.convert import ResourceTree, StringArray, Plurals, xml2po, Translation
from babel.plural import _plural_tags as PLURAL_TAGS
from babel.messages import pofile, Message, Catalog
try:
    import cStringIO as StringIO
except ImportError:
    import StringIO

def merge_xl_content(des_ws, filepath, src_worksheet_name):
    src_wb = load_workbook(filepath, use_iterators=True)
    src_ws = src_wb.get_sheet_by_name(name = src_worksheet_name)

    for row in src_ws.iter_rows():
        values = [filepath,]
        for cell in row:
            values.append(cell.internal_value)
        des_ws.append(values)

def separate_xl_content(src_filepath):
    src_wb = load_workbook(src_filepath, use_iterators=True)
    src_ws = src_wb.get_sheet_by_name(name = "Sheet")

    mytree = {}
    for row in src_ws.iter_rows():
        subxlfilename = row[0].internal_value
        if not mytree.has_key(subxlfilename):
            mytree[subxlfilename] = []

        values = []
        for cell in row[1:]:
            values.append(cell.internal_value)

        mytree[subxlfilename].append(values)

    ret = []
    for subxlfilename in mytree.keys():
        wb = Workbook()
        ws = wb.get_sheet_by_name(name="Sheet")

        for values in mytree[subxlfilename]:
            ws.append(values)

        wb.save(subxlfilename)

        ret.append(subxlfilename)

    return ret

def isXlsxFile(path):
    return match(r'(.)*xlsx$', path)

def merge_component_xl(dir, xl_name):
    def visit_xl_file(workbook, dirname, names):
        for name in names:
            filepath = os.path.join(dirname, name)
            print "filepath: %s" % filepath
            print 'match: %s' % isXlsxFile(filepath)

            if os.path.isfile(filepath) and isXlsxFile(filepath):
                merge_xl_content(workbook, filepath, 'Sheet')

    result = Workbook()
    ws = result.get_sheet_by_name(name=r"Sheet")

    os.path.walk(dir, visit_xl_file, ws)

    result.save(xl_name)

def upodate_component_xl_content2pofile(component_xl_file):
    component_dir = os.path.dirname(component_xl_file)

    # TODO: delete all po files.
    po_dict = {}

    src_wb = load_workbook(component_xl_file, use_iterators=True)
    src_ws = src_wb.get_sheet_by_name(name='Sheet')

    for row in src_ws.iter_rows():
        pofilename = row[0].internal_value
        if not po_dict.has_key(pofilename):
            po_dict[pofilename] = []

        values = []
        for cell in row[1:]:
            values.append(cell.internal_value)

        po_dict[pofilename].append(values)

    for pofilename in po_dict.keys():
        pofilepath = os.path.join(component_dir, pofilename)
        contents = po_dict[pofilename]
        catalog = convert_xlsm_content(contents)

        with open(pofilepath, 'w') as f:
            pofile.write_po(f, catalog)

def convert_xlsm_content(contents):
    catalog = Catalog()

    for content in contents:
        resid = content[0]
        res_org_content = content[1]
        rescontent = content[2]

        if resid == None:
            continue

        if rescontent == None:
            rescontent = u""

        if match(r"\((.+)(::(.+))+\)", res_org_content):
            msg_id = res_org_content[1:-1].split("::")
            msg_string = rescontent[1:-1].split("::")
            catalog.add(msg_id, context=resid, string=msg_string)
        else:
            if match(r"\".+\"", res_org_content):
                res_org_content = res_org_content[1:-1]
            if match(r"\".+\"", rescontent):
                rescontent = rescontent[1:-1]
            catalog.add(res_org_content, context=resid, string=rescontent)

    return catalog



def test1():
    wb = Workbook()
    ws = wb.get_sheet_by_name(name=r"Sheet")

    merge_xl_content(ws, '/home/huzhennan/Works/local/Gallery2/books.xlsx', 'Sheet')

    wb.save('test.xlsx')

def test2():
    merge_component_xl('/home/huzhennan/Works/local/', 'total.xlsx')

def test_separate_xl_content():
    separate_xl_content('/home/huzhennan/totol.xlsx')

def test_upodate_component_xl_content2pofile():
    upodate_component_xl_content2pofile("/home/huzhennan/Works/local/Gallery2/books.xlsx")

if __name__ == '__main__':
    #test1()
    #test2()
    #test_separate_xl_content()
    test_upodate_component_xl_content2pofile()
